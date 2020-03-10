import os
import csv
import time
import glob
import keras
import codecs
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from keras.models import Model
from skimage import io, transform
from keras.models import Sequential
from sklearn.model_selection import StratifiedKFold
from keras.layers.convolutional import Convolution2D, MaxPooling2D, Conv2D
from keras.layers import Flatten, Dense, Input, Dropout, BatchNormalization


class History(keras.callbacks.Callback):
    def on_train_begin(self, logs={}):
        self.train_loss = {'batch': [], 'epoch': []}
        self.train_acc = {'batch': [], 'epoch': []}
        self.val_loss = {'batch': [], 'epoch': []}
        self.val_acc = {'batch': [], 'epoch': []}

    def on_batch_end(self, batch, logs={}):
        self.train_loss['batch'].append(logs.get('loss'))
        self.train_acc['batch'].append(logs.get('acc'))
        self.val_loss['batch'].append(logs.get('val_loss'))
        self.val_acc['batch'].append(logs.get('val_acc'))

    def on_epoch_end(self, batch, logs={}):
        self.train_loss['epoch'].append(logs.get('loss'))
        self.train_acc['epoch'].append(logs.get('acc'))
        self.val_loss['epoch'].append(logs.get('val_loss'))
        self.val_acc['epoch'].append(logs.get('val_acc'))

    def make_remark(self):
        wb = Workbook()
        ws = wb.create_sheet('train_acc_batch')
        ws.append(self.train_acc['batch'])
        ws = wb.create_sheet('train_loss_batch')
        ws.append(self.train_loss['batch'])
        ws = wb.create_sheet('train_acc_epoch')
        ws.append(self.train_acc['epoch'])
        ws = wb.create_sheet('train_loss_epoch')
        ws.append(self.train_loss['epoch'])
        ws = wb.create_sheet('test_acc_epoch')
        ws.append(self.val_acc['epoch'])
        ws = wb.create_sheet('test_loss_epoch')
        ws.append(self.val_loss['epoch'])
        wb.save('./excel/plot_data_VGG_Grading.xlsx')

    def loss_plot(self, loss_type):
        total_iteration = range(len(self.train_loss[loss_type]))
        plt.figure()
        plt.plot(total_iteration, self.train_acc[loss_type], color='r', marker='o',
                 markerfacecolor='red', markersize=2, label='train acc')
        plt.plot(total_iteration, self.train_loss[loss_type], color='b', marker='o',
                 markerfacecolor='blue', markersize=2, label='train loss')
        plt.xlabel(loss_type)
        plt.ylabel('train-acc-loss')
        plt.legend(loc="upper right")
        plt.savefig('./image/train-acc-loss_Grading.png')
        plt.show()
        plt.clf()

        if loss_type == 'epoch':
            plt.plot(total_iteration, self.val_acc[loss_type], color='r', marker='o',
                     markerfacecolor='red', markersize=2, label='val acc')
            plt.plot(total_iteration, self.val_loss[loss_type], color='b', marker='o',
                     markerfacecolor='blue', markersize=2, label='val loss')
            plt.xlabel(loss_type)
            plt.ylabel('val-acc-loss')
            plt.legend(loc="upper right")
            plt.savefig('./image/val_acc-loss_Grading.png')
            plt.show()

def vgg_16():
    model = Sequential()

    model.add(Convolution2D(64, (3, 3), strides=(1, 1), input_shape=(224, 224, 1), padding='same', activation='relu',
                     kernel_initializer='uniform'))
    model.add(Convolution2D(64, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Convolution2D(128, (3, 2), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(128, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Convolution2D(256, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(256, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(256, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))

    model.add(Flatten())
    model.add(Dense(4096, activation='relu'))
    model.add(Dropout(0.5))
    model.add(Dense(4096, activation='relu'))
    model.add(Dropout(0.5))
    model.add(Dense(2, activation='sigmoid'))
    adam = keras.optimizers.adam(lr=0.01)
    model.compile(optimizer=adam, loss='binary_crossentropy', metrics=['accuracy'])
    return model


def vgg_13():

    model = Sequential()

    model.add(Convolution2D(64, (3, 3), strides=(1, 1), input_shape=(224, 224, 1), padding='same', activation='relu',
                     kernel_initializer='uniform'))
    model.add(Convolution2D(64, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))
    model.add(Convolution2D(128, (3, 2), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(128, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))
    model.add(Convolution2D(256, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(256, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(Convolution2D(512, (3, 3), strides=(1, 1), padding='same', activation='relu', kernel_initializer='uniform'))
    model.add(MaxPooling2D(pool_size=(2, 2)))
    model.add(Flatten())
    model.add(Dense(4096, activation='relu'))
    model.add(Dropout(0.5))
    model.add(Dense(4096, activation='relu'))
    model.add(Dropout(0.5))
    model.add(Dense(2, activation='sigmoid'))
    adam = keras.optimizers.adam(lr=0.0001)
    model.compile(optimizer=adam, loss='binary_crossentropy', metrics=['accuracy'])
    return model


def read_img(path):
    cate = [path+x for x in os.listdir(path) if os.path.isdir(path+x)]
    imgs = []
    labels = []
    for idx, folder in enumerate(cate):
        for im in glob.glob(folder+'/*.jpg'):
            print('reading the images:%s'%(im))
            img = io.imread(im)
            img = transform.resize(img, (w, h, c), mode='constant')
            imgs.append(img)
            labels.append(idx)
    return np.asarray(imgs, np.float32), np.asarray(labels, np.int32)



path = '../DATA/TrainingData_Grading/'


w = 224
h = 224
c = 1
nb_epoch = 50
batch_size = 32

data, label = read_img(path)

start_time = time.time()

# 打乱顺序
num_example = data.shape[0]
arr = np.arange(num_example)
np.random.shuffle(arr)
data = data[arr]
label = label[arr]

# 将所有数据分为训练集和验证集
seed = 7
np.random.seed(seed)
fold = 5


K_fold = StratifiedKFold(n_splits=fold, shuffle=True, random_state=seed)
loss_scores = []
acc_scores = []
round = 0
for train, test in K_fold.split(data, label):
    round += 1
    x_train = data[train]
    x_test = data[test]
    y_train = label[train]
    y_test = label[test]
    y_train = keras.utils.to_categorical(y_train, num_classes=2)
    y_test = keras.utils.to_categorical(y_test, num_classes=2)
    history = History()
    model = vgg_13()
    model.summary()
    model.fit(x_train, y_train, batch_size=batch_size, nb_epoch=nb_epoch,
              verbose=1, validation_data=(x_test, y_test), callbacks=[history])
    score = model.evaluate(x_test, y_test, verbose=0)
    print('Test score:', score[0])
    print('Test accuracy:', score[1])
    acc_scores.append(score[1])

    if round == fold:
        history.make_remark()
        history.loss_plot('epoch')
        end_time = time.time()
        duration = end_time - start_time
        print('duration:', duration)
        model.save('./model/VGG_Grading.h5')


print(acc_scores)
print(np.average(acc_scores))

